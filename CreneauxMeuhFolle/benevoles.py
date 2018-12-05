# -*- coding: utf-8 -*-
"""
Created on Tue Jun 13 15:35:43 2017

@author: GANJO
"""

import os

from PyQt5.QtCore import QSettings as Qst
import openpyxl as ox

import MessBox as M

PATH = os.getcwd()
OPTION = os.path.join(PATH,"doc","__init.ini")
HORAIRE = ["09H-10H", "10H-11H", "11H-12H", "12H-13H", "13H-14H", "14H-15H", "15H-16H", "16H-17H",
           "17H-18H", "18H-19H", "19H-20H", "20H-21H", "21H-22H", "22H-23H", "23H-00H", "00H-01H", "01H-02H", "02H-03H",
           "03H-04H", "04H-05H", "05H-06H", "06H-07H", "07H-08H", "08H-09H"]

def get_critere(creneau):
    return HORAIRE.index(creneau.split(" : ")[0])

def chaine_creneaux(creneaux_benevole):
    """A partir d'une liste de creneaux on crée
        une chaine de caratère avec saut de ligne
    """
    creneaux_benevole = sorted(creneaux_benevole, key=get_critere)
    creneaux = ''
    for k in range(len(creneaux_benevole)-1):
        creneaux += \
        creneaux_benevole[k] + "\n"
    if len(creneaux_benevole) > 0:
        creneaux += \
        creneaux_benevole[-1]
    return creneaux



class Benevoles():
    """Class représentant l'objet benevoles,
        Variable de classe :
        nombre_benevoles : Représente
    """
    nombreBenevoles = 0
    listBenevoles = []
    attribut = ["nom",
                "cre1", "cre2", "cre3",
                "artiste_vendredi", "artiste_samedi",
                "remarques"]
#,"creneaux_vendredi","creneaux_samedi","actif_vendredi","actif_samedi"

    def __init__(self, nom="",
                 cre1="", cre2="", cre3="",
                 artiste_vendredi="", artiste_samedi="",
                 remarques=""):
        if nom:
            self.nom = nom.title()
            if not cre1:
                cre1 = ''
            self.cre1 = cre1.title()
            if not cre2:
                cre2 = ''
            self.cre2 = cre2.title()
            if not cre3:
                cre3 = ''
            self.cre3 = cre3.title()
            if not artiste_vendredi:
                artiste_vendredi = ''
            self.artiste_vendredi = artiste_vendredi.title()
            if not artiste_samedi:
                artiste_samedi = ''
            self.artiste_samedi = artiste_samedi.title()
            if not remarques:
                remarques = ''

            self.remarques = remarques.title()
            self.creneaux_vendredi = []
            self.creneaux_samedi = []
            self.actif_vendredi = True
            self.actif_samedi = True

            Benevoles.listBenevoles.append(self)
            Benevoles.nombreBenevoles += 1

    @classmethod
    def set_list_benevoles(cls):
        """ Permet de charger la liste contenant tout les benevoles créés"""
        fichier = Qst(OPTION, Qst.IniFormat)
        if fichier.contains("Tableau/tab_benevoles"):
            cls.listBenevoles = fichier.value("Tableau/tab_benevoles")
            cls.nombreBenevoles = len(Benevoles.listBenevoles)

    @classmethod
    def svg_tab_benevoles(cls):
        """ Permet de sauvegarder la liste contenant tout les benevoles créés"""
        if len(cls.listBenevoles) > 0:
            fichier = Qst(OPTION, Qst.IniFormat)
            fichier.setValue("Tableau/tab_benevoles",
                             cls.listBenevoles)


    @classmethod
    def import_benevoles_excel(cls, file_name,tableau):
        """ Methode pour importer les donnees à partir d'excel
            Les Donnees Excel doivent être au bon format
            ...
        """
        cls.listBenevoles = []
        cls.nombreBenevoles = 0
        data = ox.load_workbook(file_name,
                                read_only=True,
                                keep_vba=True)
        sheet = data.get_sheet_by_name("Benevoles")
        # la première ligne est toujours l'intitulé des colonnes
        ligne = 1
        while sheet.cell(column=1, row=ligne+1).value != None:
            ligne += 1
            nom = sheet.cell(column=1, row=ligne).value
            cre1 = sheet.cell(column=2, row=ligne).value
            cre2 = sheet.cell(column=3, row=ligne).value
            cre3 = sheet.cell(column=4, row=ligne).value
            artiste_vendredi = sheet.cell(column=5, row=ligne).value
            artiste_samedi = sheet.cell(column=6, row=ligne).value
            remarques = sheet.cell(column=7, row=ligne).value
            creneaux_vendredi = sheet.cell(column=8, row=ligne).value
            creneaux_samedi = sheet.cell(column=9, row=ligne).value
            new_benevole = Benevoles(nom,
                                     cre1, cre2, cre3,
                                     artiste_vendredi, artiste_samedi,
                                     remarques)
            if creneaux_vendredi:
                new_benevole.creneaux_vendredi = creneaux_vendredi.split('\n')
            if creneaux_samedi:
                new_benevole.creneaux_samedi = creneaux_samedi.split('\n')
        sheet = data.get_sheet_by_name("Outils")
        tableau.blockSignals(True)
        for i in range(4):
            item = M.item_tab_hor_artistes(sheet.cell(column=4, row=i+2).value)
            tableau.setItem(i, 0, item)
            item = M.item_tab_hor_artistes(sheet.cell(column=5, row=i+2).value)
            tableau.setItem(i, 1, item)
        tableau.blockSignals(False)



    @classmethod
    def export_benevoles_excel(cls, file_name):
        """Methode pour exporter les resultats vers l'excel
            Les bénevoels sont remplacés si le nom du bénévoles n'est pas dans l'excel
                Sinon les valeur dans l'excel sont changés
            Les créneaux sont supprimer et les créneaux utilisés dans l'application sont ceux copiés
        """
        data = ox.load_workbook(file_name, keep_vba=True)
        sheet = data.get_sheet_by_name("Benevoles")
        nb_benevoles_application = cls.nombreBenevoles
        nb_benevoles_excel = 1
        while sheet.cell(column=1, row=nb_benevoles_excel).value != None:
            nb_benevoles_excel += 1
        cpt_benevoles = 0
        for i in range(nb_benevoles_application):
            existe = False
            for j in range(2, nb_benevoles_excel):
                benevole = cls.listBenevoles[i]
                if benevole.nom == \
                sheet.cell(column=1, row=j).value.title():
                    existe = True
                    sheet.cell(column=1, row=j,
                               value=benevole.nom)
                    sheet.cell(column=2, row=j,
                               value=benevole.cre1)
                    sheet.cell(column=3, row=j,
                               value=benevole.cre2)
                    sheet.cell(column=4, row=j,
                               value=benevole.cre3)
                    sheet.cell(column=5, row=j,
                               value=benevole.artiste_vendredi)
                    sheet.cell(column=6, row=j,
                               value=benevole.artiste_samedi)
                    sheet.cell(column=7, row=j,
                               value=benevole.remarques)
                    #### fonction de tri pour la srotie d'algo vers excel


                    creneaux_vendredi = chaine_creneaux(benevole.creneaux_vendredi)
                    sheet.cell(column=8, row=j,
                               value=creneaux_vendredi)
                    creneaux_samedi = chaine_creneaux(benevole.creneaux_samedi)
                    sheet.cell(column=9, row=j,
                               value=creneaux_samedi)

                    sheet.cell(column=10, row=j,
                               value=len(benevole.creneaux_vendredi))
                    sheet.cell(column=11, row=j,
                               value=len(benevole.creneaux_samedi))
                    break
            if not existe:
                nouvelle_ligne = nb_benevoles_excel+cpt_benevoles
                sheet.cell(column=1, row=nouvelle_ligne,
                           value=benevole.nom)
                sheet.cell(column=2, row=nouvelle_ligne,
                           value=benevole.cre1)
                sheet.cell(column=3, row=nouvelle_ligne,
                           value=benevole.cre2)
                sheet.cell(column=4, row=nouvelle_ligne,
                           value=benevole.cre3)
                sheet.cell(column=5, row=nouvelle_ligne,
                           value=benevole.artiste_vendredi)
                sheet.cell(column=6, row=nouvelle_ligne,
                           value=benevole.artiste_samedi)
                sheet.cell(column=7, row=nouvelle_ligne,
                           value=benevole.remarques)

                creneaux_vendredi = chaine_creneaux(benevole.creneaux_vendredi)
                sheet.cell(column=8, row=nouvelle_ligne,
                           value=creneaux_vendredi)
                creneaux_samedi = chaine_creneaux(benevole.creneaux_samedi)
                sheet.cell(column=9, row=nouvelle_ligne,
                           value=creneaux_samedi)
                cpt_benevoles += 1
        data.save(file_name)

    @classmethod
    def remplir_tab_benevoles(cls, tab, quantite_creneaux):
        """ Permet de remplir le tableau Benevoles dans la fenetre graphique"""
        nombre_benevoles = len(cls.listBenevoles)
        tab.setRowCount(nombre_benevoles)
        for i in range(nombre_benevoles):
            benevole = cls.listBenevoles[i]
            item = M.item_normal(benevole.nom)
            tab.setItem(i, 0, item)
            item = M.item_normal_non_editable(benevole.cre1)
            tab.setItem(i, 1, item)
            item = M.item_normal_non_editable(benevole.cre2)
            tab.setItem(i, 2, item)
            item = M.item_normal_non_editable(benevole.cre3)
            tab.setItem(i, 3, item)
            item = M.item_normal_non_editable(benevole.artiste_vendredi)
            tab.setItem(i, 4, item)
            item = M.item_normal_non_editable(benevole.artiste_samedi)
            tab.setItem(i, 5, item)
            item = M.item_normal(benevole.remarques)
            tab.setItem(i, 6, item)
        cls.maj_benevoles(tab, quantite_creneaux)

    @classmethod
    def maj_benevoles(cls, tab, quantite_creneaux):
        """Complement de la fonction  remplir_tab_benevoles(tab, quantite_creneaux)
        Permet de mettre à jour la dernière partie du tableau
        concernant le nombre de creneaux par benevole
        """
        tab.blockSignals(True)
        for i in range(len(cls.listBenevoles)):
            benevole = cls.listBenevoles[i]

            creneaux_vendredi = chaine_creneaux(benevole.creneaux_vendredi)
            item = M.item_normal_non_editable(creneaux_vendredi)
            tab.setItem(i, 7, item)

            creneaux_samedi = chaine_creneaux(benevole.creneaux_samedi)
            item = M.item_normal_non_editable(creneaux_samedi)
            tab.setItem(i, 8, item)

            nb_creneaux_vendredi = len(benevole.creneaux_vendredi)
            item = M.item_normal_nb_creneaux(nb_creneaux_vendredi,
                                             quantite_creneaux)
            tab.setItem(i, 9, item)

            nb_creneaux_samedi = len(benevole.creneaux_samedi)
            item = M.item_normal_nb_creneaux(nb_creneaux_samedi,
                                             quantite_creneaux)
            tab.setItem(i, 10, item)
        tab.resizeRowsToContents()
        tab.blockSignals(False)