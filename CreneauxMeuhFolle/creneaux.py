# -*- coding: utf-8 -*-
"""
Created on Fri Jun  9 16:45:12 2017

@author: GANJO
"""


from PyQt5.QtCore import QSettings as Qst
from PyQt5 import QtWidgets, QtCore
import openpyxl as ox
import os
import MessBox as M


PATH = os.getcwd()
OPTION = os.path.join(PATH,"doc","__init.ini")
HORAIRE = ["09H-10H", "10H-11H", "11H-12H", "12H-13H", "13H-14H", "14H-15H", "15H-16H", "16H-17H",
               "17H-18H", "18H-19H", "19H-20H", "20H-21H", "21H-22H", "22H-23H", "23H-00H", "00H-01H", "01H-02H", "02H-03H",
               "03H-04H", "04H-05H", "05H-06H", "06H-07H", "07H-08H", "08H-09H"]


class Creneaux():
    """Class représentant l'objet Creneaux,
    Variable de classe :
        nombre_creneaux : Représente
    """
    attribut = ["nom", "plage_horaire", "duree", "categorie", "benevole_vendredi", "benevole_samedi"]
    nombreCreneaux = 0
    listCreneaux = []
    catCreneaux = []
    etatCatCre = []

    def __init__(self, nom="", plage_horaire="", duree="", categorie=""):
        #On utilise le constructuer uniquement s'il s'agit bien d'un creneaux non vide
        if nom:
            self.nom = str(nom).title()
            self.plage_horaire = str(plage_horaire).title()
            self.duree = str(duree).title()
            self.categorie = str(categorie).title()
            self.loc = ['', '']
            self.benevole_vendredi = ""
            self.benevole_samedi = ""
            Creneaux.listCreneaux.append(self)
            Creneaux.nombreCreneaux += 1
            if not(self.categorie in Creneaux.catCreneaux):
                if self.categorie != "" and Creneaux.nombreCreneaux != 1:
                    Creneaux.catCreneaux.append(self.categorie)
                    Creneaux.etatCatCre.append('2')

    @classmethod
    def set_list_creneaux(cls):
        """ Permet de charger la liste contenant tout les creneaux déjà créés"""
        fichier = Qst(OPTION, Qst.IniFormat)
        if fichier.contains("Tableau/tab_creneau"):
            cls.listCreneaux = fichier.value("Tableau/tab_creneau")
            cls.nombreCreneaux = len(cls.listCreneaux)
        if fichier.contains("List/categorie_Creneau"):
            cls.catCreneaux = fichier.value("List/categorie_Creneau")
            cls.etatCatCre = fichier.value("List/categorie_Creneau_state")

    @classmethod
    def svg_list_creneaux(cls, listwid):
        """ Permet de sauvegarder la liste contenant tout les creneaux déjà créés"""
        if len(cls.listCreneaux) > 0:
            fichier = Qst(OPTION, Qst.IniFormat)
            fichier.setValue("Tableau/tab_creneau", cls.listCreneaux)
            nbVal = listwid.count() #nb valeur dans la liste
            val = []
            state = []
            if nbVal > 0:
                for i in range(nbVal):
                    val.append(listwid.item(i).text().title()) #enregistrement des valeurs
                    state.append(str(listwid.item(i).checkState())) #si checkable enregistrement des etat 0 Unchecked , 2 Checked
                fichier.setValue("List/categorie_Creneau_state",state)
                fichier.setValue("List/categorie_Creneau",val)

    @classmethod
    def charger_categorie_creneau(cls, listwid):#fonction pour charger les donnees d'une Qlist
        dejaPresent = []
        if cls.catCreneaux:
            for i in range(listwid.count()):
                dejaPresent.append(listwid.item(i).text().title())
            for i in range(len(cls.catCreneaux)):
                if not(cls.catCreneaux[i] in dejaPresent) :
                    item = QtWidgets.QListWidgetItem(cls.catCreneaux[i].title(), listwid)
                    item.setFlags(item.flags() | QtCore.Qt.ItemIsUserCheckable)
                    if int(cls.etatCatCre[i]) == 2:
                        item.setCheckState(QtCore.Qt.Checked)
                    if int(cls.etatCatCre[i]) == 0:
                        item.setCheckState(QtCore.Qt.Unchecked)
                    item.setTextAlignment(QtCore.Qt.AlignCenter)

    @classmethod
    def maj_categorie_creneau(cls, listwid):
        nbVal = listwid.count() #nb valeur dans la liste
        cls.catCreneaux = []
        cls.etatCatCre = []
        if nbVal > 0:
            for i in range(nbVal):
                cls.catCreneaux.append(listwid.item(i).text().title()) #enregistrement des valeurs
                cls.etatCatCre.append(str(listwid.item(i).checkState())) #si checkable enregistrement des etat 0 Unchecked , 2 Checked
            listwid.item(nbVal-1).setTextAlignment(QtCore.Qt.AlignCenter)

    def get_critere(creneau):
        return HORAIRE.index(creneau.plage_horaire), creneau.categorie, creneau.duree, creneau.nom

    @classmethod
    def tri_creneaux(cls):
        cls.listCreneaux = sorted(cls.listCreneaux, key=cls.get_critere)

    def get_critere_auto(creneau):
        return creneau.duree

    @classmethod
    def tri_creneaux_auto(cls):
        cls.listCreneaux = sorted(cls.listCreneaux, key=cls.get_critere_auto, reverse=True)

    @classmethod
    def import_creneaux_excel(cls, file_name):
        """ Methode pour importer les donnees à partir d'excel
            Les Donnees Excel doivent être au bon format
            ...
        """
        cls.listCreneaux = []
        cls.catCreneaux = []
        cls.etatCatCre = []
        cls.nombreCreneaux = 0
        data = ox.load_workbook(file_name, read_only=True, keep_vba=True)
        sheet = data.get_sheet_by_name("Creneaux")
        # la première ligne est toujours l'intitulé des colonnes
        ligne = 1
        while sheet.cell(column=1, row=ligne+1).value != None:
            ligne += 1
            nom = sheet.cell(column=1, row=ligne).value
            plage_horaire = sheet.cell(column=2, row=ligne).value
            duree = sheet.cell(column=3, row=ligne).value
            categorie = sheet.cell(column=4, row=ligne).value
            benevole_vendredi = sheet.cell(column=5, row=ligne).value
            benevole_samedi = sheet.cell(column=6, row=ligne).value

            creneaux = Creneaux(nom, plage_horaire, duree, categorie)

            if benevole_vendredi:
                creneaux.benevole_vendredi = benevole_vendredi.title()
            if benevole_samedi:
                creneaux.benevole_samedi = benevole_samedi.title()

    @classmethod
    def export_creneaux_excel(cls, file_name):
        """Methode pour exporter les resultats vers l'excel
            Les bénevoels sont remplacés si le nom du bénévoles n'est pas dans l'excel
                Sinon les valeur dans l'excel sont changés
            Les créneaux sont supprimer et les créneaux utilisés dans l'application sont ceux copiés
        """
        data = ox.load_workbook(file_name, keep_vba=True)
        sheet = data.get_sheet_by_name("Creneaux")
        ligne = 1
        while sheet.cell(column=1, row=ligne).value != None:
            ligne += 1

        for i in range(cls.nombreCreneaux):
            creneau = cls.listCreneaux[i]
            sheet.cell(column=1, row=i+2, value=creneau.nom)
            sheet.cell(column=2, row=i+2, value=creneau.plage_horaire)
            sheet.cell(column=3, row=i+2, value=creneau.duree)
            sheet.cell(column=4, row=i+2, value=creneau.categorie)
            sheet.cell(column=5, row=i+2, value=creneau.benevole_vendredi)
            sheet.cell(column=6, row=i+2, value=creneau.benevole_samedi)

        for i in range(cls.nombreCreneaux+2, ligne):
            for j in range(1, 7):
                sheet.cell(column=j, row=i).value = None
        data.save(file_name)

    @classmethod
    def remplir_tab_creneau(cls, tab):
        """ Permet de remplir le tableau creneaux dans la fenetre graphique"""
        tab.setRowCount(cls.nombreCreneaux)
        for i in range(cls.nombreCreneaux):
            creneau = cls.listCreneaux[i]
            item = M.item_normal(creneau.nom)
            tab.setItem(i, 0, item)
            item = M.item_normal_non_editable(creneau.plage_horaire)
            tab.setItem(i, 1, item)
            item = M.item_normal_non_editable(creneau.duree)
            tab.setItem(i, 2, item)
            item = M.item_normal_non_editable(creneau.categorie)
            tab.setItem(i, 3, item)
        cls.maj_creneaux(tab)

    @classmethod
    def maj_creneaux(cls, tab):
        """Complement de la fonction  remplir_tab_benevoles(tab, quantite_creneaux)
        Permet de mettre à jour la dernière partie du tableau
        concernant le benevoles pour chaque creneau
        """
        tab.blockSignals(True)
        for i in range(cls.nombreCreneaux):
            creneau = cls.listCreneaux[i]
            item = M.item_normal_non_editable(creneau.benevole_vendredi)
            tab.setItem(i, 4, item)
            item = M.item_normal_non_editable(creneau.benevole_samedi)
            tab.setItem(i, 5, item)
        tab.blockSignals(False)
